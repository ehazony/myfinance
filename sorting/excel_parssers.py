from abc import abstractmethod

import openpyxl

from myFinance.models import TransactionNameTag
from datetime import datetime


class _ExcelParser:
    @abstractmethod
    def load_transactions(self, ws, user):
        pass

    @abstractmethod
    def is_right_format(self, ws):
        pass


class CalExcelParser(_ExcelParser):

    def load_transactions(self, ws, user):
        sorted_transactions = list()
        unsorted_transactions = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in list(ws.iter_rows())[3:-1]:
            if not row[1].value or not row[3].value or not row[0].value:
                continue
            row_data = {}
            row_data["name"] = row[1].value
            row_data["value"] = row[3].value.replace("₪", "").replace(",", "").strip()
            row_data["date"] = row[0].value
            tag = TransactionNameTag.get_tag(row_data["name"], user)
            if tag:
                row_data["tag"] = tag.name
                sorted_transactions.append(row_data)
            else:
                unsorted_transactions.append(row_data)
        return sorted_transactions + unsorted_transactions

    def is_right_format(self, ws):
        lines = list(ws.iter_rows())
        A1 = "פירוט עסקות נכון לתאריך"
        A2 = "פירוט עסקות לכרטיס ויזה עסקי"
        titels = ["תאריך העסקה", "שם בית העסק", "סכום העסקה", "סכום החיוב", "פירוט נוסף"]
        return [x.value for x in lines[2]] == titels


class MaxExcelParser(_ExcelParser):

    def load_transactions(self, ws, user):
        sorted_transactions = list()
        unsorted_transactions = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in list(ws.iter_rows())[4:-1]:
            if not row[1].value or not row[5].value or not row[0].value:
                continue
            row_data = {}
            row_data["name"] = row[1].value
            row_data["value"] = row[5].value
            row_data["date"] = datetime.strptime(row[0].value, '%d-%m-%Y')
            tag = TransactionNameTag.get_tag(row_data["name"], user)
            if tag:
                row_data["tag"] = tag.name
                sorted_transactions.append(row_data)
            else:
                unsorted_transactions.append(row_data)
        return sorted_transactions + unsorted_transactions

    def is_right_format(self, ws):
        lines = list(ws.iter_rows())

        A1 = "פירוט עסקות נכון לתאריך"
        A2 = "פירוט עסקות לכרטיס ויזה עסקי"
        titels = ['תאריך עסקה',
                  'שם בית העסק',
                  'קטגוריה',
                  '4 ספרות אחרונות של כרטיס האשראי',
                  'סוג עסקה',
                  'סכום חיוב',
                  'מטבע חיוב',
                  'סכום עסקה מקורי',
                  'מטבע עסקה מקורי',
                  'תאריך חיוב',
                  'הערות',
                  'מועדון הנחות',
                  'מפתח דיסקונט',
                  'אופן ביצוע ההעסקה',
                  'שער המרה ממטבע מקור/התחשבנות לש"ח']
        return [x.value for x in lines[3]] == titels


class DiscountBankExcelParser(_ExcelParser):

    def load_transactions(self, ws, user):
        pass

    def is_right_format(self, ws):
        titels = ['תאריך',
                  'יום ערך',
                  'תיאור התנועה',
                  '₪ זכות/חובה ',
                  '₪ יתרה ',
                  'אסמכתה',
                  'עמלה',
                  'ערוץ ביצוע']
        lines = list(ws.iter_rows())
        return [x.value for x in lines[12]] == titels





def load_excel_file(ws, user):
    sorted_transactions = list()
    unsorted_transactions = list()
    # iterating over the rows and
    # getting value from each cell in row
    for row in list(ws.iter_rows()):
        if not row[1].value or not row[3].value or not row[0].value:
            continue
        row_data = {}
        row_data["name"] = row[1].value
        row_data["value"] = row[3].value.replace("₪", "").replace(",", "").strip()
        row_data["date"] = row[0].value
        tag = TransactionNameTag.get_tag(row_data["name"], user)
        if tag:
            row_data["tag"] = tag.name
            sorted_transactions.append(row_data)
        else:
            unsorted_transactions.append(row_data)
    return sorted_transactions + unsorted_transactions


class ExcelParser:
    parsers = [CalExcelParser(), MaxExcelParser(), DiscountBankExcelParser()]

    def parse_excel(self, excel_file, user):
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.worksheets[0]
        for parser in self.parsers:
            if parser.is_right_format(worksheet):
                return parser.load_transactions(worksheet, user)
        raise ValueError('excel not recegnised')
